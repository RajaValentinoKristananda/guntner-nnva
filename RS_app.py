import os
import json
import base64
import openpyxl
from datetime import datetime, timedelta
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import google.generativeai as genai
from PIL import Image, ImageDraw, ImageFont
import io

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PT Guntner — NNVA System",
    page_icon="assets/güntner_logo.png" if os.path.exists("assets/güntner_logo.png") else None,
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Paths ─────────────────────────────────────────────────────────────────────
DATA_FOLDER   = 'data'
ASSETS_FOLDER = 'assets'
os.makedirs(DATA_FOLDER,   exist_ok=True)
os.makedirs(ASSETS_FOLDER, exist_ok=True)

PO_DATA_FILE         = os.path.join(DATA_FOLDER, 'po_data.xlsx')
RECOMMENDATIONS_FILE = os.path.join(DATA_FOLDER, 'recommendations.xlsx')
KNOWLEDGE_FILE       = os.path.join(DATA_FOLDER, 'company_knowledge.xlsx')
AI_INSIGHTS_FILE     = os.path.join(DATA_FOLDER, 'ai_insights.xlsx')
LAYOUT_EXISTING_FILE = os.path.join(DATA_FOLDER, 'layout_p1_existing.pdf')
LAYOUT_BLANK_FILE    = os.path.join(DATA_FOLDER, 'layout_p1_blank.pdf')
LAYOUT_MAPS_FOLDER   = os.path.join(DATA_FOLDER, 'layout_maps')
os.makedirs(LAYOUT_MAPS_FOLDER, exist_ok=True)

# ── Gemini ────────────────────────────────────────────────────────────────────
GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
genai.configure(api_key=GEMINI_API_KEY)
gemini = genai.GenerativeModel('gemini-2.5-flash')

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
#MainMenu, footer { visibility: hidden; }
html, body, [class*="css"] {
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    -webkit-font-smoothing: antialiased;
}
.main { background-color: #f8fafc; }
[data-testid="stHeader"] { background: transparent !important; }
.block-container {
    padding-top: 2rem !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    padding-bottom: 2rem !important;
    max-width: 100% !important;
}
.stTabs [data-baseweb="tab-list"] {
    gap: 0.5rem; background-color: white;
    border-radius: 0.75rem; padding: 0.5rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
}
.stTabs [data-baseweb="tab"] {
    background-color: transparent; border-radius: 0.5rem;
    padding: 0.75rem 1.5rem; font-weight: 600;
    color: #64748b; border: none;
}
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    background-color: #3b82f6; color: white;
}
.kpi-card {
    background: white; border-radius: 14px;
    border: 1px solid #e2e8f0; padding: 1.5rem 1.5rem 1.25rem;
    position: relative; overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04); transition: box-shadow .2s;
}
.kpi-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.08); }
.kpi-card::before {
    content: ''; position: absolute;
    top: 0; left: 0; right: 0; height: 3px;
    border-radius: 14px 14px 0 0;
}
.kpi-card.blue::before   { background: #3b82f6; }
.kpi-card.violet::before { background: #8b5cf6; }
.kpi-card.green::before  { background: #10b981; }
.kpi-card.amber::before  { background: #f59e0b; }
.kpi-label {
    font-size: 0.7rem; font-weight: 700;
    letter-spacing: 0.09em; text-transform: uppercase;
    color: #94a3b8; margin-bottom: 0.65rem;
}
.kpi-value {
    font-size: 2.4rem; font-weight: 700;
    color: #0f172a; letter-spacing: -0.04em;
    line-height: 1; font-family: 'DM Mono', monospace;
}
.kpi-unit { font-size: 0.8rem; color: #94a3b8; margin-top: 4px; }
.ai-banner {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    border-radius: 14px; padding: 1.5rem 1.75rem; margin: 1.25rem 0;
    display: flex; gap: 1.25rem; align-items: flex-start;
    box-shadow: 0 4px 16px rgba(102,126,234,0.35);
}
.ai-banner-badge {
    font-size: 0.62rem; font-weight: 700;
    letter-spacing: 0.12em; text-transform: uppercase;
    color: rgba(255,255,255,0.75); margin-bottom: 0.5rem;
}
.ai-banner-text { font-size: 0.9rem; line-height: 1.7; color: rgba(255,255,255,0.92); }
.strip-empty {
    background: #f8fafc; border: 1.5px dashed #cbd5e1;
    border-radius: 10px; padding: 1rem 1.25rem;
    color: #64748b; font-size: 0.875rem;
    display: flex; align-items: center; gap: 0.6rem; margin: 0.75rem 0;
}
.strip-info {
    background: #eff6ff; border-left: 3px solid #3b82f6;
    border-radius: 8px; padding: 0.85rem 1.25rem;
    color: #1e40af; font-size: 0.875rem; margin: 0.75rem 0;
}
.strip-success {
    background: #f0fdf4; border-left: 3px solid #10b981;
    border-radius: 8px; padding: 0.85rem 1.25rem;
    color: #166534; font-size: 0.875rem; font-weight: 500; margin: 0.75rem 0;
}
.sec-label {
    font-size: 0.68rem; font-weight: 700;
    letter-spacing: 0.1em; text-transform: uppercase;
    color: #94a3b8; margin: 1.75rem 0 0.75rem;
}
.stButton > button {
    background: #3b82f6; color: #ffffff; border: none;
    border-radius: 8px; padding: 0.55rem 1.35rem;
    font-weight: 600; font-size: 0.85rem;
    font-family: 'DM Sans', sans-serif;
    transition: background .18s, transform .1s;
}
.stButton > button:hover { background: #2563eb; transform: translateY(-1px); }
.stSelectbox > div > div,
.stTextInput > div > div > input,
.stDateInput > div > div > input,
.stTextArea > div > div > textarea {
    border-radius: 8px; border-color: #e2e8f0;
    font-family: 'DM Sans', sans-serif;
    font-size: 0.875rem; background: #ffffff;
}
[data-testid="stDataFrame"] {
    border-radius: 10px; overflow: hidden; border: 1px solid #e2e8f0;
}
.stExpander { border: 1px solid #e2e8f0 !important; border-radius: 10px !important; background: white; }
[data-testid="metric-container"] {
    background: white; border: 1px solid #e2e8f0;
    border-radius: 12px; padding: 1rem 1.25rem;
}
</style>
""", unsafe_allow_html=True)

# ── Logo header ───────────────────────────────────────────────────────────────
_logo_html = ''
if os.path.exists('assets/güntner_logo.png'):
    with open('assets/güntner_logo.png', 'rb') as f:
        _logo_b64 = base64.b64encode(f.read()).decode()
    _logo_html = f'<img src="data:image/png;base64,{_logo_b64}" style="height:72px;">'

st.markdown(
    f'<div style="background:white;border-bottom:1px solid #e5e7eb;padding:14px 28px;'
    f'display:flex;align-items:center;gap:16px;margin:-2rem -2.5rem 1.5rem -2.5rem;">'
    f'{_logo_html}'
    f'<div>'
    f'<div style="font-size:32px;font-weight:700;color:#111827;line-height:1.2;'
    f'font-family:Arial,sans-serif;">Cycle Time Dashboard</div>'
    f'<div style="font-size:16px;color:#9ca3af;margin-top:2px;">'
    f'Monitoring Non-Value Added activities and progress towards lean goals</div>'
    f'</div></div>',
    unsafe_allow_html=True
)

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_dash, tab_upload, tab_ai, tab_history, tab_data = st.tabs([
    "Dashboard", "Upload Data", "AI Analysis", "Recommendations History", "Data Manager"
])

# ── Plotly theme ──────────────────────────────────────────────────────────────
def _theme(fig, height=380):
    fig.update_layout(
        paper_bgcolor='white', plot_bgcolor='white',
        font=dict(family='DM Sans, sans-serif', color='#374151', size=12),
        height=height, margin=dict(l=8, r=8, t=36, b=8),
        xaxis=dict(showgrid=False, linecolor='#e2e8f0', tickfont=dict(size=11)),
        yaxis=dict(showgrid=True, gridcolor='#f1f5f9', linecolor='#e2e8f0', tickfont=dict(size=11)),
        legend=dict(bgcolor='rgba(0,0,0,0)', font=dict(size=11)),
    )
    return fig

# ── Init Excel ────────────────────────────────────────────────────────────────
def init_excel_files():
    if not os.path.exists(PO_DATA_FILE):
        pd.DataFrame(columns=[
            'po_no','process_section','time_sec','time_type',
            'vz_sec','te_sec','recording_date','area','item_no','created_at'
        ]).to_excel(PO_DATA_FILE, index=False, engine='openpyxl')
    if not os.path.exists(RECOMMENDATIONS_FILE):
        pd.DataFrame(columns=[
            'id','week_start','week_end','process_section','total_nnva_time',
            'occurrence_count','analysis','recommendations_json','layout_map_path','created_at'
        ]).to_excel(RECOMMENDATIONS_FILE, index=False, engine='openpyxl')
    if not os.path.exists(AI_INSIGHTS_FILE):
        pd.DataFrame(columns=['id','month','year','insight_text','created_at']
        ).to_excel(AI_INSIGHTS_FILE, index=False, engine='openpyxl')
    if not os.path.exists(KNOWLEDGE_FILE):
        pd.DataFrame({
            'category': ['company_profile','manufacturing_processes','machines','lean_principles'],
            'content': [
                'PT Guntner Indonesia — fabrikasi komponen logam presisi untuk otomotif & elektronik.',
                'Proses: Welding, Bending, Punching, Assembly, Quality Control.',
                'Mesin: Robot Welding, Press Brake, CNC Punching, Hydraulic Press, CMM.',
                'Lean: eliminasi waste (muda, mura, muri). VA = value-add, NVA = hapus, NNVA = optimasi.',
            ]
        }).to_excel(KNOWLEDGE_FILE, index=False, engine='openpyxl')

init_excel_files()


def get_layout_bytes():
    parts = []
    for path in [LAYOUT_EXISTING_FILE, LAYOUT_BLANK_FILE]:
        if os.path.exists(path):
            with open(path, 'rb') as f:
                parts.append(f.read())
    return parts


# ── PDF to image (PyMuPDF) ────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def pdf_to_image_bytes(pdf_bytes):
    try:
        import fitz
        doc  = fitz.open(stream=pdf_bytes, filetype="pdf")
        page = doc[0]
        mat  = fitz.Matrix(2.5, 2.5)
        pix  = page.get_pixmap(matrix=mat)
        return pix.tobytes("png")
    except Exception as e:
        st.error(f"PDF conversion error: {e}")
        return None


# ── Layout annotation renderer ────────────────────────────────────────────────
def render_layout_annotation(layout_img_bytes, zones):
    img = Image.open(io.BytesIO(layout_img_bytes)).convert('RGBA')
    W, H = img.size
    overlay = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw    = ImageDraw.Draw(overlay)

    COLORS = {
        'problem':   ((220, 38,  38,  90), (220, 38,  38,  255)),
        'solution':  ((16,  185, 129, 80), (16,  185, 129, 255)),
        'flow':      ((59,  130, 246, 65), (59,  130, 246, 255)),
        'secondary': ((245, 158, 11,  75), (245, 158, 11,  255)),
    }

    TYPE_NAME = {
        'problem':   '❌ MASALAH',
        'solution':  '✅ SOLUSI',
        'flow':      '➡️ JALUR',
        'secondary': '⚠️ DAMPAK',
    }

    try:
        font_big   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22)
        font_mid   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 17)
        font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
        font_num   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
    except:
        font_big = font_mid = font_small = font_num = ImageFont.load_default()

    legend_items = []

    for idx, z in enumerate(zones, 1):
        ztype        = normalize_zone_type(z.get('type', 'problem'))
        fill, border = COLORS.get(ztype, COLORS['problem'])
        x1 = max(0, int(z['x1'] * W))
        y1 = max(0, int(z['y1'] * H))
        x2 = min(W, int(z['x2'] * W))
        y2 = min(H, int(z['y2'] * H))
        if x2 <= x1 or y2 <= y1:
            continue

        # Fill zona
        draw.rectangle([x1, y1, x2, y2], fill=fill)
        for t in range(4):
            draw.rectangle([x1-t, y1-t, x2+t, y2+t], outline=border)

        # ── Label box di dalam zona ───────────────────────────────────
        action_text = z.get('action', z.get('label', ''))[:50]
        label_text  = z.get('label', '')[:35]

        # Background putih untuk label supaya mudah dibaca
        box_pad = 6
        box_h   = 58
        box_y1  = y1 + 4
        box_y2  = box_y1 + box_h
        box_x1  = x1 + 4
        box_x2  = min(x2 - 4, x1 + 420)

        draw.rectangle([box_x1, box_y1, box_x2, box_y2],
                       fill=(255, 255, 255, 210), outline=border)

        # Nomor zona di kotak kiri
        num_cx = box_x1 + 18
        num_cy = box_y1 + 29
        draw.ellipse([num_cx-16, num_cy-16, num_cx+16, num_cy+16],
                     fill=border, outline=(255,255,255,255))
        draw.text((num_cx, num_cy), str(idx),
                  fill=(255,255,255,255), font=font_num, anchor='mm')

        # Teks label (aksi)
        draw.text((box_x1 + 42, box_y1 + box_pad),
                  label_text, fill=(20,20,20,255), font=font_mid)
        # Sub-teks action
        draw.text((box_x1 + 42, box_y1 + box_pad + 24),
                  action_text[:55], fill=(60,60,60,255), font=font_small)

        legend_items.append((idx, ztype, label_text, z.get('action',''), z.get('distance_m','')))

    # ── Legend di bawah ──────────────────────────────────────────────
    leg_row_h = 38
    leg_x     = 10
    leg_y     = H - 16 - len(legend_items) * leg_row_h - 10
    leg_w     = 680
    leg_h     = len(legend_items) * leg_row_h + 20

    draw.rectangle([leg_x - 4, leg_y - 8, leg_x + leg_w, leg_y + leg_h],
                   fill=(255,255,255,235), outline=(150,150,150,255))

    # Header legend
    draw.text((leg_x + 8, leg_y - 4), "RINGKASAN TINDAKAN:",
              fill=(30,30,30,255), font=font_mid)

    _, border_prob = COLORS['problem']
    _, border_sol  = COLORS['solution']
    _, border_flow = COLORS['flow']
    _, border_sec  = COLORS['secondary']
    color_map = {
        'problem': border_prob, 'solution': border_sol,
        'flow': border_flow, 'secondary': border_sec
    }

    for i, (num, ztype, label, action, dist) in enumerate(legend_items):
        lx  = leg_x + 8
        ly  = leg_y + 18 + i * leg_row_h
        col = color_map.get(ztype, (100,100,100,255))

        # Kotak warna kecil
        draw.rectangle([lx, ly + 4, lx + 26, ly + 26], fill=col, outline=(255,255,255,255))
        draw.text((lx + 14, ly + 16), str(num),
                  fill=(255,255,255,255), font=font_small, anchor='mm')

        # Tipe
        type_short = TYPE_NAME.get(ztype, ztype)
        draw.text((lx + 34, ly + 2), type_short,
                  fill=col, font=font_small)

        # Label aksi
        dist_str = f" [{dist}m]" if dist else ""
        draw.text((lx + 34, ly + 18),
                  (label + dist_str)[:70], fill=(30,30,30,255), font=font_small)

    composite = Image.alpha_composite(img, overlay).convert('RGB')
    buf = io.BytesIO()
    composite.save(buf, format='PNG', quality=95)
    return buf.getvalue()


VALID_ZONE_TYPES = {"problem", "solution", "flow", "secondary"}

TYPE_ALIAS = {
    "path": "flow", "jalur": "flow", "route": "flow", "corridor": "flow",
    "affected": "secondary", "terdampak": "secondary", "dampak": "secondary",
    "secondary_area": "secondary", "impact": "secondary",
    "improvement": "solution", "perbaikan": "solution", "solusi": "solution",
    "recommendation": "solution", "target": "solution",
    "masalah": "problem", "issue": "problem", "waste": "problem",
    "bottleneck": "problem", "current": "problem",
}

def normalize_zone_type(raw: str) -> str:
    t = str(raw).strip().lower()
    if t in VALID_ZONE_TYPES:
        return t
    for alias, valid in TYPE_ALIAS.items():
        if alias in t:
            return valid
    return "secondary"

def gen_layout_zones(rec, process_name, layout_bytes, ctx):
    recs_text = "\n".join(
        ["- " + r.get("judul","") + ": " + r.get("area_terdampak","")
         for r in rec.get("rekomendasi", [])]
    )

    zones_prompt = (
        "Kamu adalah sistem anotasi peta pabrik PT Guntner Plant 1.\n"
        "Layout pabrik dilampirkan sebagai PDF — BACA dengan teliti semua teks label area.\n\n"

        "=== LANGKAH WAJIB SEBELUM MENENTUKAN KOORDINAT ===\n"
        "LANGKAH 1 — IDENTIFIKASI NAMA AREA:\n"
        "  Baca semua teks yang tertulis di layout PDF.\n"
        "  Cari area/workstation yang namanya paling relevan dengan proses: '" + process_name + "'\n"
        "  Contoh: jika proses = 'melepas laminasi inside tray',\n"
        "    → cari teks seperti 'MANUAL TRAY', 'TRAY PROCESSING', 'ASSEMBLING TRAY', dll\n"
        "    → bukan asal tebak tengah layout!\n\n"

        "LANGKAH 2 — TENTUKAN KOORDINAT DARI NAMA AREA YANG DITEMUKAN:\n"
        "  Setelah menemukan nama area tersebut di PDF,\n"
        "  baru tentukan koordinat x1,y1,x2,y2 berdasarkan POSISI ASLI teks/kotak area itu.\n"
        "  Koordinat adalah RASIO 0.0-1.0 dari lebar/tinggi halaman PDF.\n\n"

        "=== ATURAN ZONA ===\n"
        "Buat TEPAT 4 zona:\n\n"

        "ZONA 1 — type: 'problem' (WAJIB: lokasi area yang namanya relevan dengan proses NNVA)\n"
        "  → Ini HARUS di area yang namanya kamu temukan di langkah 1\n"
        "  → Ukuran: lebar ~0.08-0.14, tinggi ~0.07-0.12\n\n"

        "ZONA 2 — type: 'solution' (area terdekat dari zona 1 yang bisa jadi lokasi perbaikan)\n"
        "  → Pilih area nyata di sebelah/dekat zona 1, bukan tempat kosong\n"
        "  → Ukuran: lebar ~0.08-0.14, tinggi ~0.07-0.12\n\n"

        "ZONA 3 — type: 'flow' (jalur/koridor yang menghubungkan zona 1 ke zona 2)\n"
        "  → Gambarkan sebagai garis sempit memanjang\n"
        "  → Ukuran: lebar ~0.02-0.05, tinggi ~0.10-0.30\n\n"

        "ZONA 4 — type: 'secondary' (area lain yang ikut terdampak NNVA ini)\n"
        "  → Pilih area nyata di layout yang relevan\n"
        "  → Ukuran: lebar ~0.08-0.14, tinggi ~0.07-0.12\n\n"

        "=== ATURAN UKURAN (WAJIB) ===\n"
        "- Zona TIDAK BOLEH menutupi >10% area gambar\n"
        "- Lebar maksimal zona: 0.15 (kecuali flow: 0.05)\n"
        "- Tinggi maksimal zona: 0.18 (kecuali flow: 0.35)\n\n"

        "=== REKOMENDASI PERBAIKAN ===\n" + recs_text + "\n\n"

        "=== ATURAN LABEL ===\n"
        "label = nama area yang KAMU TEMUKAN DI PDF + tindakan singkat (max 35 karakter)\n"
        "  BENAR: 'Manual Tray — kurangi jarak material'\n"
        "  SALAH: 'Area bermasalah di tengah'\n"
        "action = 1 kalimat konkret tindakan perbaikan\n"
        "area_name = nama persis area seperti tertulis di PDF (wajib diisi!)\n"
        "distance_m = estimasi jarak dalam meter (angka saja, contoh: '8')\n\n"

        "Kembalikan HANYA JSON (tidak ada teks lain):\n"
        "{\n"
        '  "area_identified": "Nama area yang ditemukan di PDF untuk proses ini",\n'
        '  "zones": [\n'
        "    {\n"
        '      "label": "NamaArea — tindakan singkat",\n'
        '      "area_name": "Nama persis dari PDF",\n'
        '      "type": "problem",\n'
        '      "x1": 0.35, "y1": 0.30, "x2": 0.46, "y2": 0.40,\n'
        '      "action": "Satu kalimat tindakan konkret",\n'
        '      "description": "Kenapa area ini dipilih, berdasarkan teks di layout",\n'
        '      "distance_m": "8"\n'
        "    }\n"
        "  ]\n"
        "}\n"
        "INGAT: Zona 1 (problem) HARUS berada di area yang namanya relevan dengan '"
        + process_name + "', bukan di tempat random!"
    )

    try:
        vision   = genai.GenerativeModel("gemini-2.5-flash")
        pdf_part = {
            "inline_data": {
                "mime_type": "application/pdf",
                "data": base64.b64encode(layout_bytes).decode("utf-8")
            }
        }
        response = vision.generate_content([pdf_part, zones_prompt])
        text     = response.text

        # Debug: tampilkan area yang ditemukan Gemini
        if '"area_identified"' in text:
            try:
                raw = json.loads(text.split("```json")[-1].split("```")[0].strip()
                                 if "```" in text else text)
                area_found = raw.get("area_identified","")
                if area_found:
                    st.info(f"🔍 Gemini mengidentifikasi area: **{area_found}**")
            except:
                pass

        for fence in ["```json", "```"]:
            if fence in text:
                s    = text.find(fence) + len(fence)
                e    = text.find("```", s)
                text = text[s:e].strip()
                break

        data  = json.loads(text)
        zones = data.get("zones", [])

        fixed_zones = []
        for z in zones:
            z["type"] = normalize_zone_type(z.get("type", "secondary"))

            if not z.get("distance_m") or str(z.get("distance_m","")).strip() in ("", "—", "-"):
                default_dist = {"problem": "5", "flow": "15", "solution": "8", "secondary": "20"}
                z["distance_m"] = default_dist.get(z["type"], "10")

            x1, y1, x2, y2 = (float(z.get("x1", 0.3)), float(z.get("y1", 0.3)),
                               float(z.get("x2", 0.45)), float(z.get("y2", 0.45)))
            max_w = 0.16
            max_h = 0.18
            if z["type"] == "flow":
                max_w = 0.05
                max_h = 0.35

            if (x2 - x1) > max_w:
                cx = (x1 + x2) / 2
                x1, x2 = cx - max_w/2, cx + max_w/2
            if (y2 - y1) > max_h:
                cy = (y1 + y2) / 2
                y1, y2 = cy - max_h/2, cy + max_h/2

            # Pastikan dalam batas 0-1
            x1, y1 = max(0.01, x1), max(0.01, y1)
            x2, y2 = min(0.99, x2), min(0.99, y2)

            z.update({"x1": round(x1,3), "y1": round(y1,3),
                      "x2": round(x2,3), "y2": round(y2,3)})

            # Label fallback ke area_name jika ada
            if not z.get("label"):
                z["label"] = z.get("area_name", f"Zona {z['type']}")

            fixed_zones.append(z)

        return fixed_zones

    except Exception as e:
        st.warning(f"Zone generation error: {e}")
        return None


def get_company_context():
    try:
        df = pd.read_excel(KNOWLEDGE_FILE, engine='openpyxl')
        return "PT Guntner Indonesia:\n\n" + "\n\n".join(df['content'].tolist())
    except:
        return "PT Guntner Indonesia - Perusahaan manufaktur komponen logam presisi."


# ── Data helpers ──────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_po_data():
    try:
        df = pd.read_excel(PO_DATA_FILE, engine='openpyxl')
        for col in ['recording_date','created_at']:
            if col in df.columns: df[col] = pd.to_datetime(df[col])
        for col in ['process_section','area','item_no','po_no','time_type']:
            if col in df.columns: df[col] = df[col].fillna('')
        return df
    except:
        return pd.DataFrame(columns=[
            'po_no','process_section','time_sec','time_type',
            'vz_sec','te_sec','recording_date','area','item_no','created_at'
        ])


def save_po_data(df):
    try:
        df.to_excel(PO_DATA_FILE, index=False, engine='openpyxl')
        load_po_data.clear(); return True
    except Exception as e:
        st.error(f"Save error: {e}"); return False


def _pick(row, *names, default='', numeric=False):
    for n in names:
        if n in row.index and pd.notna(row.get(n)):
            val = row.get(n)
            if numeric:
                try: return float(val)
                except: continue
            return str(val).strip()
    return 0.0 if numeric else default


def parse_cycle_time_template(file_bytes, po_override=None):
    """Parse template Cycle Time dengan format khusus PT Guntner."""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Baca metadata dari header
    order_no    = str(ws.cell(4, 1).value or '').strip()
    area        = str(ws.cell(2, 5).value or '').strip()
    item_no     = str(ws.cell(4, 5).value or '').strip()
    rec_date_raw = ws.cell(4, 14).value
    try:
        rec_date = pd.to_datetime(rec_date_raw)
    except:
        rec_date = pd.to_datetime(datetime.now())

    po_no = po_override if po_override else (order_no or f"PO-{datetime.now().strftime('%Y%m%d')}")

    rows = []
    for r in range(9, ws.max_row + 1):
        serial = ws.cell(r, 1).value
        if serial is None:
            break  # berhenti jika serial kosong
        process = ws.cell(r, 3).value
        time_sec = ws.cell(r, 14).value
        time_type = ws.cell(r, 16).value
        vz_sec   = ws.cell(r, 18).value
        te_sec   = ws.cell(r, 20).value

        # Skip baris jika process kosong atau time_sec bukan angka
        if not process or not isinstance(time_sec, (int, float)):
            continue

        rows.append({
            'po_no':           po_no,
            'process_section': str(process).strip(),
            'time_sec':        float(time_sec),
            'time_type':       str(time_type or 'NNVA').strip().upper(),
            'vz_sec':          float(vz_sec) if isinstance(vz_sec, (int, float)) else 0.0,
            'te_sec':          float(te_sec) if isinstance(te_sec, (int, float)) else 0.0,
            'recording_date':  rec_date,
            'area':            area,
            'item_no':         item_no,
            'created_at':      pd.to_datetime(datetime.now()),
        })
    return rows, po_no, rec_date

def insert_po_data(uploaded_file_bytes, po_no_override, recording_date_override):
    """Upload dari template Cycle Time PT Guntner."""
    existing = load_po_data()
    rows, po_no, rec_date = parse_cycle_time_template(uploaded_file_bytes, po_no_override)
    if not rows:
        return 0
    combined = pd.concat([existing, pd.DataFrame(rows)], ignore_index=True)
    return len(rows) if save_po_data(combined) else 0

def insert_bulk_manual(po_no, processes_str, times_str, types_str,
                        rec_date, areas='', items='', vz='', te=''):
    existing = load_po_data()
    procs = [p.strip() for p in processes_str.split(',') if p.strip()]
    times = [t.strip() for t in times_str.split(',') if t.strip()]
    types = [t.strip() for t in types_str.split(',') if t.strip()] if types_str else []
    def _spl(s): return [x.strip() for x in s.split(',')] if s else []
    area_l, item_l, vz_l, te_l = _spl(areas), _spl(items), _spl(vz), _spl(te)
    n = len(procs)
    if len(times) != n: return 0, "Process/time count mismatch"
    for lst, d in [(types,'NNVA'),(area_l,''),(item_l,''),(vz_l,'0'),(te_l,'0')]:
        while len(lst) < n: lst.append(d)
    rows = []
    for i in range(n):
        try:
            rows.append({
                'po_no': po_no, 'process_section': procs[i],
                'time_sec': float(times[i]), 'time_type': types[i].upper(),
                'vz_sec': float(vz_l[i]) if vz_l[i] else 0,
                'te_sec': float(te_l[i]) if te_l[i] else 0,
                'recording_date': pd.to_datetime(rec_date),
                'area': area_l[i], 'item_no': item_l[i],
                'created_at': pd.to_datetime(datetime.now()),
            })
        except Exception as e: return 0, f"Error entry {i+1}: {e}"
    combined = pd.concat([existing, pd.DataFrame(rows)], ignore_index=True)
    return (len(rows),"Success") if save_po_data(combined) else (0,"Save failed")


def get_monthly_nnva(year, month):
    df = load_po_data()
    if df.empty: return pd.DataFrame()
    mask = ((df['time_type']=='NNVA') &
            (df['recording_date'].dt.year==year) &
            (df['recording_date'].dt.month==month))
    filt = df[mask]
    if filt.empty: return pd.DataFrame()
    return (filt.groupby('process_section')
                .agg(total_time=('time_sec','sum'), occurrence=('process_section','count'))
                .reset_index().sort_values('total_time',ascending=False).head(10))


# ── AI helpers ────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_ai_insights():
    try:
        df = pd.read_excel(AI_INSIGHTS_FILE, engine='openpyxl')
        if 'created_at' in df.columns: df['created_at'] = pd.to_datetime(df['created_at'])
        return df
    except: return pd.DataFrame()


def get_cached_insight(year, month):
    ins = load_ai_insights()
    if ins.empty: return None
    filt = ins[(ins['year']==year)&(ins['month']==month)]
    if filt.empty: return None
    return filt.sort_values('created_at',ascending=False).iloc[0]['insight_text']


def save_ai_insight(year, month, text):
    df = load_ai_insights()
    new_id = 1 if df.empty else int(df['id'].max())+1
    pd.concat([df,pd.DataFrame([{
        'id':new_id,'month':month,'year':year,
        'insight_text':text,'created_at':datetime.now()
    }])],ignore_index=True).to_excel(AI_INSIGHTS_FILE,index=False,engine='openpyxl')
    load_ai_insights.clear()

def gen_quick_insight(year, month, summary, ctx):
    if summary.empty: return "Tidak ada data NNVA untuk periode ini."
    top   = summary.iloc[0]
    total = summary['total_time'].sum()
    top3  = summary.head(3)
    top3_text = "\n".join([
        f"  {i+1}. {r['process_section']}: {r['total_time']:.0f}s ({r['total_time']/60:.1f} menit), {r['occurrence']}x"
        for i, (_, r) in enumerate(top3.iterrows())
    ])
    prompt = (
        f"{ctx}\n\n"
        f"DATA NNVA {month}/{year}:\n"
        f"- Total NNVA: {total:.0f}s ({total/60:.1f} menit)\n"
        f"- Proses tertinggi: {top['process_section']} — {top['total_time']:.0f}s ({top['total_time']/60:.1f} menit)\n"
        f"- Frekuensi: {top['occurrence']} kali\n\n"
        "Tulis 2-3 kalimat dalam bahasa Indonesia profesional.\n"
        "Kalimat 1: sebutkan proses NNVA tertinggi, waktu yang terbuang, dan frekuensinya.\n"
        "Kalimat 2: jelaskan kemungkinan penyebab NNVA-nya (misal: jarak material jauh, "
        "tidak ada alat bantu, metode tidak standar, dll).\n"
        "Kalimat 3: berikan 1 saran konkret paling mudah diterapkan untuk mengurangi waktu NNVA tersebut "
        "(contoh: redesain tata letak, tambah jig/fixture, standarisasi SOP, relocate material, dll).\n"
        "Jangan sebut proses lain. Tidak ada bullet, tidak ada judul, tidak ada bold. Langsung paragraf."
    )
    try: return gemini.generate_content(prompt).text.strip()
    except Exception as e: return f"Error: {e}"


@st.cache_data(ttl=300)
def load_recommendations():
    try:
        df = pd.read_excel(RECOMMENDATIONS_FILE, engine='openpyxl')
        if 'created_at' in df.columns: df['created_at'] = pd.to_datetime(df['created_at'])
        if 'layout_map_path' not in df.columns: df['layout_map_path'] = ''
        return df
    except: return pd.DataFrame()


def save_recommendation(ws, we, proc, total_t, occ, analysis, rec_json, map_path=''):
    df     = load_recommendations()
    new_id = 1 if df.empty else int(df['id'].max())+1
    if 'layout_map_path' not in df.columns: df['layout_map_path'] = ''
    pd.concat([df, pd.DataFrame([{
        'id': new_id, 'week_start': ws, 'week_end': we,
        'process_section': proc, 'total_nnva_time': total_t,
        'occurrence_count': occ, 'analysis': analysis,
        'recommendations_json': json.dumps(rec_json, ensure_ascii=False),
        'layout_map_path': map_path,
        'created_at': datetime.now()
    }])], ignore_index=True).to_excel(RECOMMENDATIONS_FILE, index=False, engine='openpyxl')
    load_recommendations.clear()
    return new_id

def update_recommendation_map(rec_id, map_path, zones=None):
    """Update layout_map_path dan zones untuk record tertentu."""
    df = load_recommendations()
    if df.empty: return
    if 'layout_map_path' not in df.columns:
        df['layout_map_path'] = ''
    if 'layout_map_zones' not in df.columns:
        df['layout_map_zones'] = ''
    df.loc[df['id'] == rec_id, 'layout_map_path'] = map_path
    if zones:
        df.loc[df['id'] == rec_id, 'layout_map_zones'] = json.dumps(zones, ensure_ascii=False)
    df.to_excel(RECOMMENDATIONS_FILE, index=False, engine='openpyxl')
    load_recommendations.clear()

def save_map_to_file(map_bytes, rec_id, process_name):
    safe_name = "".join(c if c.isalnum() else "_" for c in process_name)[:30]
    filename  = f"map_{rec_id}_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
    filepath  = os.path.join(LAYOUT_MAPS_FOLDER, filename)
    with open(filepath, 'wb') as f:
        f.write(map_bytes)
    return filepath

def gen_recommendation(proc, time_sec, occ, ctx, all_processes=None, layout_bytes=None):
    all_proc_text = ""
    if all_processes is not None and not all_processes.empty:
        rows = ["  - {}: {:.0f}s ({} kali)".format(
            r["process_section"], r["total_time"], r["occurrence"]
        ) for _, r in all_processes.iterrows()]
        all_proc_text = "\nSEMUA PROSES NNVA:\n" + "\n".join(rows)

    layout_ctx = ""
    if layout_bytes:
        layout_ctx = ("\n\nLAYOUT PABRIK: PDF layout Plant 1 PT Guntner Indonesia dilampirkan. "
                      "Pelajari posisi area, jalur material flow, dan jarak antar area. "
                      "Gunakan untuk saran SPESIFIK — sebutkan nama area/section nyata.")

    prompt_text = (
        f"{ctx}{layout_ctx}\n\n"
        f"FOKUS ANALISIS:\nProses: {proc}\n"
        f"Total NNVA: {time_sec:.2f}s ({time_sec/60:.1f} menit)\n"
        f"Frekuensi: {occ} kali\n{all_proc_text}\n\n"
        "Berikan analisis mendalam dengan 3 rekomendasi SPESIFIK.\n\n"
        "Output JSON valid (tidak ada teks lain di luar JSON):\n"
        '{"layout_analysis":"Analisis layout: area yang relevan, jarak material flow, bottleneck fisik",'
        '"analisis":"Analisis mendalam mengapa NNVA",'
        '"root_cause":"Akar penyebab dikaitkan dengan layout dan alur kerja",'
        '"rekomendasi":['
        '{"judul":"...","deskripsi":"...","area_terdampak":"...","dampak":"...","kesulitan":"Mudah/Sedang/Sulit","biaya":"Rendah/Sedang/Tinggi","roi_estimate":"...","langkah":["..."],"tools_needed":["..."],"timeline":"..."},'
        '{"judul":"...","deskripsi":"...","area_terdampak":"...","dampak":"...","kesulitan":"...","biaya":"...","roi_estimate":"...","langkah":["..."],"tools_needed":["..."],"timeline":"..."},'
        '{"judul":"...","deskripsi":"...","area_terdampak":"...","dampak":"...","kesulitan":"...","biaya":"...","roi_estimate":"...","langkah":["..."],"tools_needed":["..."],"timeline":"..."}'
        '],"prioritas_implementasi":"...","quick_wins":"..."}'
    )

    try:
        if layout_bytes:
            vision   = genai.GenerativeModel("gemini-2.5-flash")
            pdf_part = {"inline_data":{"mime_type":"application/pdf","data":base64.b64encode(layout_bytes).decode("utf-8")}}
            response = vision.generate_content([pdf_part, prompt_text])
        else:
            response = gemini.generate_content(prompt_text)

        text = response.text
        for fence in ["```json", "```"]:
            if fence in text:
                s = text.find(fence) + len(fence)
                e = text.find("```", s)
                text = text[s:e].strip()
                break
        return json.loads(text)
    except Exception as e:
        st.error(f"AI error: {e}"); return None


# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
with tab_dash:
    po_data = load_po_data()

    if po_data.empty:
        st.markdown('<div class="strip-empty">⚠️ No data yet — upload PO data first.</div>',
                    unsafe_allow_html=True)
    else:
        MONTHS = ["January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
        cur_yr = datetime.now().year
        cur_mo = datetime.now().month

        _s, fc1, fc2, fc3 = st.columns([4,2,2,1])
        with fc1:
            years  = list(range(cur_yr-2,cur_yr+2))
            sel_yr = st.selectbox("Year", years, index=years.index(cur_yr))
        with fc2:
            sel_mo_name = st.selectbox("Month", MONTHS, index=cur_mo-1)
            sel_mo = MONTHS.index(sel_mo_name)+1
        with fc3:
            st.write(""); st.write("")
            if st.button("↺"):
                load_po_data.clear(); load_ai_insights.clear(); st.rerun()

        monthly_sum = get_monthly_nnva(sel_yr, sel_mo)

        st.markdown("## Performance Overview")
        k1,k2,k3,k4 = st.columns(4)
        kpis = [
            (k1,"Total Records",   len(po_data),                              "",    "blue"),
            (k2,"NNVA Time",       f"{po_data[po_data['time_type']=='NNVA']['time_sec'].sum()/60:.1f}", "min","violet"),
            (k3,"Processes",       po_data['process_section'].nunique(),       "",    "green"),
            (k4,"Total POs",       po_data['po_no'].nunique(),                 "",    "amber"),
        ]
        for col,label,val,unit,color in kpis:
            with col:
                unit_html = f'<div class="kpi-unit">{unit}</div>' if unit else ''
                st.markdown(f'<div class="kpi-card {color}"><div class="kpi-label">{label}</div>'
                            f'<div class="kpi-value">{val}</div>{unit_html}</div>',
                            unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("## Trend Analysis")

        ch1, ch2 = st.columns(2)
        with ch1:
            time_dist = po_data.groupby('time_type')['time_sec'].sum().reset_index()
            fig = px.pie(time_dist, values='time_sec', names='time_type',
                         color='time_type',
                         color_discrete_map={'VA':'#3b82f6','NNVA':'#ef4444','NVA':'#f59e0b'},
                         hole=0.5, title="Time Type Distribution")
            fig.update_traces(textposition='inside', textinfo='percent+label',
                              marker=dict(line=dict(color='white',width=2)))
            _theme(fig)
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar':False})

        with ch2:
            if not monthly_sum.empty:
                fig2 = px.bar(monthly_sum, x='total_time', y='process_section',
                              orientation='h', color='total_time',
                              color_continuous_scale=[[0,'#fecaca'],[.5,'#f87171'],[1,'#dc2626']],
                              labels={'total_time':'Time (sec)','process_section':''},
                              title=f"Top NNVA Processes — {sel_mo_name}")
                fig2.update_layout(coloraxis_showscale=False,
                                   yaxis={'categoryorder':'total ascending'})
                _theme(fig2)
                st.plotly_chart(fig2, use_container_width=True, config={'displayModeBar':False})
            else:
                st.markdown(f'<div class="strip-empty">📭 No NNVA data for {sel_mo_name} {sel_yr}.</div>',
                            unsafe_allow_html=True)

        monthly_trend = (po_data[po_data['time_type']=='NNVA']
                         .assign(Month=lambda d: d['recording_date'].dt.to_period('M'))
                         .groupby('Month',as_index=False)['time_sec'].sum()
                         .sort_values('Month'))
        if not monthly_trend.empty:
            monthly_trend['Month_str'] = monthly_trend['Month'].astype(str)
            fig3 = go.Figure()
            fig3.add_trace(go.Scatter(
                x=monthly_trend['Month_str'], y=monthly_trend['time_sec'],
                mode='lines', fill='tozeroy', line=dict(color='#ef4444',width=0),
                fillcolor='rgba(239,68,68,0.08)', hoverinfo='skip', showlegend=False
            ))
            fig3.add_trace(go.Scatter(
                x=monthly_trend['Month_str'], y=monthly_trend['time_sec'],
                mode='lines+markers', line=dict(color='#ef4444',width=3),
                marker=dict(size=7,color='#ef4444',line=dict(width=2,color='white')),
                hovertemplate='<b>%{x}</b><br>%{y:,.0f} sec<extra></extra>', showlegend=False
            ))
            fig3.update_layout(title='Monthly NNVA Time Trend')
            _theme(fig3, height=280)
            st.plotly_chart(fig3, use_container_width=True, config={'displayModeBar':False})

        cached = get_cached_insight(sel_yr, sel_mo)
        if cached:
            ins_col1, ins_col2 = st.columns([9, 1])
            with ins_col1:
                st.markdown(f'''<div class="ai-banner">
                    <div style="font-size:1.4rem;flex-shrink:0;margin-top:2px;">✦</div>
                    <div>
                        <div class="ai-banner-badge">AI Insight — {sel_mo_name} {sel_yr}</div>
                        <div class="ai-banner-text">{cached}</div>
                    </div>
                </div>''', unsafe_allow_html=True)
            with ins_col2:
                st.write("")
                if st.button("🗑️", key="del_insight", help="Hapus insight ini untuk generate ulang"):
                    df_ins = load_ai_insights()
                    df_ins = df_ins[~((df_ins['year']==sel_yr) & (df_ins['month']==sel_mo))]
                    df_ins.to_excel(AI_INSIGHTS_FILE, index=False, engine='openpyxl')
                    load_ai_insights.clear()
                    st.rerun()
        else:
            ia1, ia2 = st.columns([4,1])
            with ia1:
                st.markdown('<div class="strip-empty">✦ No AI insight yet for this period.</div>',
                            unsafe_allow_html=True)
            with ia2:
                if st.button("Generate Insight", key="gen_ins"):
                    if not monthly_sum.empty:
                        with st.spinner("AI analysing…"):
                            ctx = get_company_context()
                            ins = gen_quick_insight(sel_yr, sel_mo, monthly_sum, ctx)
                            save_ai_insight(sel_yr, sel_mo, ins)
                            st.rerun()
                    else: st.warning("No NNVA data this month.")

        st.markdown("---")
        t1, t2 = st.columns(2)
        with t1:
            st.markdown(f'<div class="sec-label">Top NNVA — {sel_mo_name} {sel_yr}</div>',
                        unsafe_allow_html=True)
            if not monthly_sum.empty:
                disp = monthly_sum.copy()
                disp['Time (min)'] = (disp['total_time']/60).round(2)
                disp = disp.rename(columns={
                    'process_section':'Process','total_time':'Time (sec)','occurrence':'Freq'
                })[['Process','Time (sec)','Time (min)','Freq']]
                st.dataframe(disp, use_container_width=True, hide_index=True)
            else:
                st.markdown('<div class="strip-empty">No records.</div>',unsafe_allow_html=True)
        with t2:
            st.markdown('<div class="sec-label">Recent Records</div>', unsafe_allow_html=True)
            recent = po_data.sort_values('created_at',ascending=False).head(10)
            show_cols = [c for c in ['po_no','process_section','time_sec','time_type','recording_date']
                         if c in recent.columns]
            st.dataframe(recent[show_cols], use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — UPLOAD DATA
# ════════════════════════════════════════════════════════════════════════════
with tab_upload:
    st.markdown("## Upload PO Data")
    ut1, ut2 = st.tabs(["Excel Upload", "Manual Entry"])

    with ut1:
        u1, u2 = st.columns([3,1])
        with u1:
            st.markdown('<div class="sec-label">File</div>', unsafe_allow_html=True)
            uploaded = st.file_uploader("Excel file", type=['xlsx','xls'],
                                        label_visibility="collapsed")
            up_po = st.text_input(
                "PO Number (kosongkan untuk ambil otomatis dari file)",
                value="",
                placeholder="Otomatis dari file",
                key="up_po")
            up_date = st.date_input("Recording Date", value=datetime.now(), key="up_date")
            if st.button("Upload", type="primary"):
                if uploaded:
                    try:
                        with st.spinner("Processing…"):
                            file_bytes = uploaded.read()
                            po_override = up_po.strip() if up_po.strip() else None
                            n = insert_po_data(file_bytes, po_override, up_date)
                        if n > 0:
                            st.markdown(f'<div class="strip-success">✓ {n} rows uploaded.</div>',
                                        unsafe_allow_html=True)
                        else:
                            st.error("Upload failed — pastikan format template Cycle Time PT Guntner.")
                    except Exception as e:
                        st.error(f"Error: {e}")
                else:
                    st.warning("Select a file first.")
        with u2:
            st.markdown('<div class="sec-label">Column Guide</div>', unsafe_allow_html=True)
            st.markdown('''<div class="strip-info">
            <strong>Required</strong><br>Process Section · Time (sec) · Time Type
            <br><br><strong>Optional</strong><br>Area · Item No · VZ, Sec · te, Sec
            </div>''', unsafe_allow_html=True)
            po_up = load_po_data()
            st.metric("Records", len(po_up))
            st.metric("POs", po_up['po_no'].nunique() if not po_up.empty else 0)

    with ut2:
        st.markdown('''<div class="strip-info">
        <strong>Bulk Entry:</strong> Comma-separated values.
        </div>''', unsafe_allow_html=True)
        m1, m2 = st.columns(2)
        with m1:
            m_po   = st.text_input("PO Number", value=f"PO-{datetime.now().strftime('%Y%m%d')}", key="m_po")
            st.markdown('<div class="sec-label">Process Sections</div>', unsafe_allow_html=True)
            m_proc = st.text_area("P", placeholder="Welding, Bending, Punching",
                                   height=90, label_visibility="collapsed")
            st.markdown('<div class="sec-label">Time in seconds</div>', unsafe_allow_html=True)
            m_time = st.text_area("T", placeholder="120.5, 80, 95",
                                   height=90, label_visibility="collapsed")
            st.markdown('<div class="sec-label">Time Types (optional)</div>', unsafe_allow_html=True)
            m_type = st.text_area("TY", placeholder="NNVA, VA, NNVA",
                                   height=70, label_visibility="collapsed")
        with m2:
            m_date = st.date_input("Recording Date", value=datetime.now(), key="m_date")
            st.markdown('<div class="sec-label">Optional Fields</div>', unsafe_allow_html=True)
            m_area = st.text_input("Area", placeholder="Production Floor A, B")
            m_item = st.text_input("Item No", placeholder="ITEM-001, ITEM-002")
            m_vz   = st.text_input("VZ, Sec", placeholder="10, 15")
            m_te   = st.text_input("TE, Sec", placeholder="5, 8")

        if m_proc and m_time:
            procs = [p.strip() for p in m_proc.split(',') if p.strip()]
            times = [t.strip() for t in m_time.split(',') if t.strip()]
            if len(procs)==len(times):
                st.markdown(f'<div class="strip-success">Ready to add {len(procs)} entries.</div>',
                            unsafe_allow_html=True)
                st.dataframe(
                    pd.DataFrame([{'#':i+1,'Process':p,'Time (sec)':t}
                                   for i,(p,t) in enumerate(zip(procs,times))]),
                    use_container_width=True, hide_index=True)
            else: st.error(f"Mismatch: {len(procs)} processes vs {len(times)} times.")

        if st.button("Add Entries", type="primary", use_container_width=True, key="m_add"):
            if m_proc and m_time:
                n, msg = insert_bulk_manual(m_po, m_proc, m_time, m_type,
                                             m_date, m_area, m_item, m_vz, m_te)
                if n > 0:
                    st.markdown(f'<div class="strip-success">✓ {n} entries added.</div>',
                                unsafe_allow_html=True)
                    st.rerun()
                else: st.error(f"Failed: {msg}")
            else: st.warning("Fill in Process Sections and Time.")

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — AI ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
with tab_ai:
    st.markdown("## AI-Powered NNVA Analysis")
    sub_layout, sub_analysis = st.tabs(["📐 Factory Layout Setup", "🔍 Run Analysis"])

    # ════ SUB-TAB 1: Layout Setup ════════════════════════════════════════
    with sub_layout:
        st.markdown("### Factory Layout Knowledge Base")

        lc1, lc2 = st.columns(2)
        with lc1:
            st.markdown('<div class="sec-label">Layout P1 Existing</div>',
                        unsafe_allow_html=True)
            existing_up = st.file_uploader("Layout Existing", type=["pdf"],
                                           label_visibility="collapsed", key="upload_existing")
            if existing_up:
                if st.button("Save Layout Existing", type="primary", key="save_existing"):
                    with open(LAYOUT_EXISTING_FILE, 'wb') as f:
                        f.write(existing_up.read())
                    st.markdown(f'<div class="strip-success">✓ Saved ({os.path.getsize(LAYOUT_EXISTING_FILE)//1024} KB)</div>',
                                unsafe_allow_html=True)
            if os.path.exists(LAYOUT_EXISTING_FILE):
                st.markdown(f'<div class="strip-success">✓ Layout Existing tersimpan — {os.path.getsize(LAYOUT_EXISTING_FILE)//1024} KB</div>',
                            unsafe_allow_html=True)
                if st.button("🗑️ Hapus Layout Existing", key="del_existing"):
                    os.remove(LAYOUT_EXISTING_FILE)
                    st.rerun()

        with lc2:
            st.markdown('<div class="sec-label">Layout Blank Plant 1</div>', unsafe_allow_html=True)
            blank_up = st.file_uploader("Layout Blank", type=["pdf"],
                                        label_visibility="collapsed", key="upload_blank")
            if blank_up:
                if st.button("Save Layout Blank", type="primary", key="save_blank"):
                    with open(LAYOUT_BLANK_FILE, 'wb') as f:
                        f.write(blank_up.read())
                    st.markdown(f'<div class="strip-success">✓ Saved ({os.path.getsize(LAYOUT_BLANK_FILE)//1024} KB)</div>',
                                unsafe_allow_html=True)
            if os.path.exists(LAYOUT_BLANK_FILE):
                st.markdown(f'<div class="strip-success">✓ Layout Blank tersimpan — {os.path.getsize(LAYOUT_BLANK_FILE)//1024} KB</div>',
                            unsafe_allow_html=True)
                if st.button("🗑️ Hapus Layout Blank", key="del_blank"):
                    os.remove(LAYOUT_BLANK_FILE)
                    st.rerun()

        st.markdown("---")
        sc1, sc2, sc3 = st.columns(3)
        layouts_ready = []
        if os.path.exists(LAYOUT_EXISTING_FILE): layouts_ready.append("Existing")
        if os.path.exists(LAYOUT_BLANK_FILE):    layouts_ready.append("Blank")
        sc1.metric("Layouts Saved", f"{len(layouts_ready)}/2")
        sc2.metric("Status", "Ready" if layouts_ready else "Not Set")
        sc3.metric("AI Mode", "Layout-Aware" if layouts_ready else "Generic")

        if layouts_ready:
            st.markdown(
                '<div class="ai-banner">'
                '<div style="font-size:1.4rem;flex-shrink:0;">🗺️</div>'
                '<div><div class="ai-banner-badge">Layout Knowledge Ready</div>'
                '<div class="ai-banner-text">Gemini akan membaca layout Plant 1 dan memberikan '
                'rekomendasi dengan estimasi jarak antar area yang presisi.</div>'
                '</div></div>', unsafe_allow_html=True)
            with st.expander("📐 Preview Layout Existing", expanded=False):
                with open(LAYOUT_EXISTING_FILE, 'rb') as f:
                    b64 = base64.b64encode(f.read()).decode()
                st.markdown(
                    f'<iframe src="data:application/pdf;base64,{b64}" '
                    f'width="100%" height="600px" '
                    f'style="border:1px solid #e2e8f0;border-radius:8px;"></iframe>',
                    unsafe_allow_html=True)
        else:
            st.markdown('<div class="strip-empty">⚠️ Belum ada layout tersimpan.</div>',
                        unsafe_allow_html=True)

    # ════ SUB-TAB 2: Run Analysis ════════════════════════════════════════
    with sub_analysis:
        layouts_saved = get_layout_bytes()
        if layouts_saved:
            st.markdown(
                f'<div class="strip-success">🗺️ {len(layouts_saved)} layout PDF tersimpan — analisis layout-aware.</div>',
                unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="strip-empty">⚠️ Belum ada layout. Upload di tab Factory Layout Setup.</div>',
                unsafe_allow_html=True)

        st.markdown("---")
        a1, a2 = st.columns([3, 1])
        today_ai = datetime.now().date()

        with a1:
            st.markdown('<div class="sec-label">Date Range</div>', unsafe_allow_html=True)
            dc1, dc2 = st.columns(2)
            with dc1: start_d = st.date_input("From", today_ai - timedelta(days=7), key="ai_start")
            with dc2: end_d   = st.date_input("To",   today_ai, key="ai_end")

            analyse_mode = st.radio(
                "Analyse scope",
                ["Top NNVA Process Only", "All NNVA Processes (comprehensive)"],
                horizontal=True
            )

            btn_label = "🗺️ Run Layout-Aware Analysis" if layouts_saved else "🔍 Run Analysis (No Layout)"
            run_btn   = st.button(btn_label, type="primary", use_container_width=True)

            if run_btn:
                with st.spinner("Gemini sedang membaca layout dan menganalisis NNVA…"):
                    po = load_po_data()
                    if po.empty:
                        st.warning("No PO data available.")
                    else:
                        mask = (
                            (po["time_type"] == "NNVA") &
                            (po["recording_date"] >= pd.to_datetime(start_d)) &
                            (po["recording_date"] <= pd.to_datetime(end_d))
                        )
                        filt = po[mask]
                        if filt.empty:
                            st.warning("No NNVA records in selected range.")
                        else:
                            summ = (filt.groupby("process_section")
                                        .agg(total_time=("time_sec","sum"),
                                             occurrence=("process_section","count"))
                                        .reset_index()
                                        .sort_values("total_time", ascending=False))
                            top_row   = summ.iloc[0]
                            ctx       = get_company_context()
                            all_procs = summ if "All" in analyse_mode else None

                            rec = gen_recommendation(
                                top_row["process_section"], top_row["total_time"],
                                top_row["occurrence"], ctx,
                                all_processes=all_procs,
                                layout_bytes=layouts_saved[0] if layouts_saved else None
                            )

                            if rec:
                                new_rec_id = save_recommendation(
                                    start_d, end_d, top_row["process_section"],
                                    top_row["total_time"], top_row["occurrence"],
                                    rec.get("analisis",""), rec
                                )
                                st.session_state["last_rec"]    = rec
                                st.session_state["last_top"]    = {
                                    "process_section": top_row["process_section"],
                                    "total_time":      float(top_row["total_time"]),
                                    "occurrence":      int(top_row["occurrence"])
                                }
                                st.session_state["last_rec_id"] = new_rec_id
                                st.session_state.pop("annotated_map", None)
                                st.session_state.pop("map_zones", None)
                                st.session_state.pop("map_saved_path", None)

            # ── Render hasil ──────────────────────────────────────────────
            if "last_rec" in st.session_state:
                rec = st.session_state["last_rec"]
                top = st.session_state["last_top"]

                st.markdown(
                    '<div class="strip-success">✓ Analisis selesai — tersimpan di Recommendations History.</div>',
                    unsafe_allow_html=True)
                st.markdown("---")

                if rec.get("layout_analysis"):
                    st.markdown(
                        '<div class="ai-banner">'
                        '<div style="font-size:1.4rem;flex-shrink:0;">🗺️</div>'
                        '<div><div class="ai-banner-badge">Gemini — Pemahaman Layout Pabrik</div>'
                        f'<div class="ai-banner-text">{rec["layout_analysis"]}</div>'
                        '</div></div>', unsafe_allow_html=True)

                ca, cb = st.columns(2)
                with ca:
                    with st.expander("📊 Analisis", expanded=True):
                        st.write(rec.get("analisis","—"))
                with cb:
                    with st.expander("🔍 Root Cause", expanded=True):
                        st.write(rec.get("root_cause","—"))

                st.markdown('<div class="sec-label">Rekomendasi Improvement</div>', unsafe_allow_html=True)
                diff_color = {"Mudah":"#10b981","Sedang":"#f59e0b","Sulit":"#ef4444"}
                for i, r in enumerate(rec.get("rekomendasi",[]),1):
                    diff = r.get("kesulitan","—")
                    dc   = diff_color.get(diff,"#64748b")
                    area = r.get("area_terdampak","")
                    with st.expander(f"{i}. {r.get('judul','—')}", expanded=True):
                        badges = ""
                        if area:
                            badges += (f'<span style="background:#eff6ff;color:#1d4ed8;font-size:0.72rem;'
                                       f'font-weight:700;padding:3px 10px;border-radius:20px;'
                                       f'border:1px solid #bfdbfe;margin-right:8px;">📍 {area}</span>')
                        badges += (f'<span style="background:{dc}22;color:{dc};font-size:0.72rem;'
                                   f'font-weight:700;padding:3px 10px;border-radius:20px;'
                                   f'border:1px solid {dc}44;">{diff}</span>')
                        st.markdown(f'<div style="margin-bottom:12px;">{badges}</div>', unsafe_allow_html=True)
                        rc1, rc2 = st.columns(2)
                        with rc1:
                            st.markdown(f"**Deskripsi**\n\n{r.get('deskripsi','—')}")
                            st.markdown(f"**Dampak:** {r.get('dampak','—')}")
                            st.markdown(f"**Timeline:** {r.get('timeline','—')}")
                        with rc2:
                            st.markdown(f"**Kesulitan:** <span style='color:{dc};font-weight:700;'>{diff}</span>",
                                        unsafe_allow_html=True)
                            st.markdown(f"**Biaya:** {r.get('biaya','—')}")
                            st.markdown(f"**ROI:** {r.get('roi_estimate','—')}")
                        st.markdown("**Langkah Implementasi**")
                        for step in r.get("langkah",[]): st.markdown(f"- {step}")
                        st.markdown("**Tools yang Dibutuhkan**")
                        for tool in r.get("tools_needed",[]): st.markdown(f"- {tool}")

                qw1, qw2 = st.columns(2)
                with qw1:
                    with st.expander("⚡ Quick Wins", expanded=True):
                        st.write(rec.get("quick_wins","—"))
                with qw2:
                    with st.expander("📋 Prioritas Implementasi"):
                        st.write(rec.get("prioritas_implementasi","—"))

                # ── Visual Layout Annotation ──────────────────────────────
                if layouts_saved:
                    st.markdown("---")
                    st.markdown('<div class="sec-label">🗺️ Visual Layout Annotation</div>',
                                unsafe_allow_html=True)
                    st.markdown(
                        '<div class="strip-info">'
                        '<strong>Merah</strong>=sumber NNVA · '
                        '<strong>Hijau</strong>=lokasi perbaikan · '
                        '<strong>Biru</strong>=jalur material · '
                        '<strong>Amber</strong>=area terdampak · '
                        'Angka di zona = estimasi jarak dalam meter.'
                        '</div>', unsafe_allow_html=True)

                    if st.button("🗺️ Generate Layout Map", type="primary", key="gen_map"):
                        with st.spinner("Gemini menentukan posisi zona secara presisi…"):
                            ctx   = get_company_context()
                            zones = gen_layout_zones(
                                rec, top["process_section"], layouts_saved[0], ctx)
                            if zones:
                                img_bytes = pdf_to_image_bytes(layouts_saved[0])
                                if img_bytes:
                                    annotated = render_layout_annotation(img_bytes, zones)
                                    st.session_state["annotated_map"] = annotated
                                    st.session_state["map_zones"]     = zones

                                    # ✅ Simpan ke file & update history (termasuk zones)
                                    rec_id   = st.session_state.get("last_rec_id")
                                    if rec_id:
                                        map_path = save_map_to_file(
                                            annotated, rec_id, top["process_section"])
                                        update_recommendation_map(rec_id, map_path, zones=zones)
                                        st.session_state["map_saved_path"] = map_path
                                else:
                                    st.error("Gagal konversi PDF ke gambar.")
                            else:
                                st.error("Gemini tidak dapat menentukan koordinat zona.")

                    if "annotated_map" in st.session_state:
                        # ── Side by side layout asli vs annotasi ──────────
                        st.markdown('<div class="sec-label">Perbandingan Layout</div>',
                                    unsafe_allow_html=True)
                        col_orig, col_ann = st.columns(2)
                        with col_orig:
                            st.markdown("**📄 Layout Asli**")
                            with open(LAYOUT_EXISTING_FILE, 'rb') as f:
                                b64_orig = base64.b64encode(f.read()).decode()
                            st.markdown(
                                f'<iframe src="data:application/pdf;base64,{b64_orig}" '
                                f'width="100%" height="500px" '
                                f'style="border:1px solid #e2e8f0;border-radius:8px;"></iframe>',
                                unsafe_allow_html=True)
                        with col_ann:
                            st.markdown("**🗺️ Layout + Annotasi NNVA**")
                            st.image(st.session_state["annotated_map"], use_container_width=True)

                        # Full size
                        st.markdown('<div class="sec-label">Annotasi Full Size</div>',
                                    unsafe_allow_html=True)
                        st.image(st.session_state["annotated_map"], use_container_width=True,
                                 caption="Layout Plant 1 — Annotasi NNVA & Rekomendasi")

                        # Keterangan zona
                        if "map_zones" in st.session_state:
                            st.markdown('<div class="sec-label">Keterangan Zona</div>',
                                        unsafe_allow_html=True)
                            type_labels = {
                                "problem":   "🔴 Area Bermasalah",
                                "solution":  "🟢 Lokasi Perbaikan",
                                "flow":      "🔵 Jalur Material",
                                "secondary": "🟡 Area Terdampak"
                            }
                            zone_rows = []
                            for i, z in enumerate(st.session_state["map_zones"],1):
                                zone_rows.append({
                                    "No":         i,
                                    "Tipe":       type_labels.get(z.get("type",""), z.get("type","")),
                                    "Area/Label": z.get("label",""),
                                    "Tindakan":   z.get("action","—"),   # ← BARU
                                    "Jarak (m)":  z.get("distance_m","—"),
                                    "Keterangan": z.get("description","")
                                })
                            st.dataframe(pd.DataFrame(zone_rows),
                                         use_container_width=True, hide_index=True)

                        dl1, dl2 = st.columns(2)
                        with dl1:
                            st.download_button(
                                label="⬇️ Download Layout Map",
                                data=st.session_state["annotated_map"],
                                file_name=f"layout_annotation_{datetime.now().strftime('%Y%m%d_%H%M')}.png",
                                mime="image/png"
                            )
                        with dl2:
                            if st.session_state.get("map_saved_path"):
                                st.markdown(
                                    '<div class="strip-success">✓ Map sudah tersimpan di History</div>',
                                    unsafe_allow_html=True)

        with a2:
            st.markdown('<div class="sec-label">Period Summary</div>', unsafe_allow_html=True)
            po2 = load_po_data()
            if not po2.empty:
                mask2 = (
                    (po2["time_type"] == "NNVA") &
                    (po2["recording_date"] >= pd.to_datetime(start_d)) &
                    (po2["recording_date"] <= pd.to_datetime(end_d))
                )
                filt2 = po2[mask2]
                if not filt2.empty:
                    s2 = (filt2.groupby("process_section")
                               .agg(total_time=("time_sec","sum"),
                                    occurrence=("process_section","count"))
                               .reset_index()
                               .sort_values("total_time", ascending=False))
                    st.metric("Processes",   len(s2))
                    st.metric("Peak Time",   f"{s2.iloc[0]['total_time']:.0f}s")
                    st.metric("Top Process", s2.iloc[0]["process_section"][:24])
                    st.markdown('<div class="sec-label">All NNVA Processes</div>',
                                unsafe_allow_html=True)
                    st.dataframe(
                        s2.rename(columns={"process_section":"Process",
                                           "total_time":"Time (s)","occurrence":"Freq"}
                        )[["Process","Time (s)","Freq"]],
                        use_container_width=True, hide_index=True)
                else:
                    st.markdown('<div class="strip-empty">No data.</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — RECOMMENDATIONS HISTORY
# ════════════════════════════════════════════════════════════════════════════
with tab_history:
    st.markdown("## Recommendations History")
    recs = load_recommendations()

    if recs.empty:
        st.markdown('<div class="strip-empty">No recommendations yet — run AI Analysis first.</div>',
                    unsafe_allow_html=True)
    else:
        recs = recs.sort_values('created_at', ascending=False)
        st.markdown(f'<div class="sec-label">{len(recs)} records</div>', unsafe_allow_html=True)

        top_procs = recs.groupby('process_section')['total_nnva_time'].sum().nlargest(10).reset_index()
        if not top_procs.empty:
            fig_h = px.bar(top_procs, x='total_nnva_time', y='process_section',
                           orientation='h', color='total_nnva_time',
                           color_continuous_scale=[[0,'#fecaca'],[1,'#dc2626']],
                           labels={'total_nnva_time':'Total NNVA (sec)','process_section':''},
                           title="Cumulative NNVA Time by Process")
            fig_h.update_layout(coloraxis_showscale=False, yaxis={'categoryorder':'total ascending'})
            _theme(fig_h, height=320)
            st.plotly_chart(fig_h, use_container_width=True, config={'displayModeBar':False})

        st.markdown("---")
        for _, row in recs.iterrows():
            map_path = str(row.get('layout_map_path',''))
            has_map  = (map_path.strip() != '' and os.path.exists(map_path))
            map_icon = " 🗺️" if has_map else ""
            lbl = (f"{row['week_start']} → {row['week_end']}  ·  "
                   f"{row['process_section']}  ·  {row['total_nnva_time']:.0f}s{map_icon}")

            with st.expander(lbl):
                hc1, hc2, hc3 = st.columns(3)
                hc1.metric("NNVA Time",   f"{row['total_nnva_time']:.1f}s")
                hc2.metric("Occurrences", row['occurrence_count'])
                hc3.metric("Created",     row['created_at'].strftime('%Y-%m-%d')
                                          if pd.notna(row['created_at']) else '—')
                st.markdown("---")

                try:
                    data = json.loads(row['recommendations_json'])
                    st.markdown(f"**Analisis:** {data.get('analisis','—')}")
                    st.markdown(f"**Root Cause:** {data.get('root_cause','—')}")
                    for i, r in enumerate(data.get('rekomendasi',[]),1):
                        st.markdown(f"**{i}. {r.get('judul','—')}** "
                                    f"— {r.get('kesulitan','?')} | {r.get('biaya','?')} biaya")
                        st.write(f"→ {r.get('deskripsi','—')}")
                except:
                    st.write(row.get('analysis','—'))

                # ── Layout Map di History ──────────────────────────────────
                if has_map:
                    st.markdown("---")
                    st.markdown('<div class="sec-label">🗺️ Layout Annotation Map</div>',
                                unsafe_allow_html=True)
                    with open(map_path, 'rb') as f:
                        map_data = f.read()

                    if os.path.exists(LAYOUT_EXISTING_FILE):
                        h1, h2 = st.columns(2)
                        with h1:
                            st.markdown("**📄 Layout Asli**")
                            with open(LAYOUT_EXISTING_FILE, 'rb') as f:
                                b64_h = base64.b64encode(f.read()).decode()
                            st.markdown(
                                f'<iframe src="data:application/pdf;base64,{b64_h}" '
                                f'width="100%" height="380px" '
                                f'style="border:1px solid #e2e8f0;border-radius:8px;"></iframe>',
                                unsafe_allow_html=True)
                        with h2:
                            st.markdown("**🗺️ Layout + Annotasi NNVA**")
                            st.image(map_data, use_container_width=True)
                    else:
                        st.image(map_data, use_container_width=True)

                    # ── Keterangan Zona di History ─────────────────────────
                    zones_json = str(row.get('layout_map_zones', ''))
                    if zones_json.strip() and zones_json.strip() != 'nan':
                        try:
                            saved_zones = json.loads(zones_json)
                            st.markdown('<div class="sec-label">Keterangan Zona</div>',
                                        unsafe_allow_html=True)
                            type_labels = {
                                "problem":   "🔴 Area Bermasalah",
                                "solution":  "🟢 Lokasi Perbaikan",
                                "flow":      "🔵 Jalur Material",
                                "secondary": "🟡 Area Terdampak"
                            }
                            zone_rows = []
                            for zi, z in enumerate(saved_zones, 1):
                                zone_rows.append({
                                    "No":         zi,
                                    "Tipe":       type_labels.get(z.get("type",""), z.get("type","")),
                                    "Area":       z.get("label",""),
                                    "Jarak (m)":  z.get("distance_m","—"),
                                    "Keterangan": z.get("description","")
                                })
                            st.dataframe(pd.DataFrame(zone_rows),
                                         use_container_width=True, hide_index=True)
                        except:
                            pass

                    st.download_button(
                        label="⬇️ Download Layout Map",
                        data=map_data,
                        file_name=f"layout_{row['process_section'][:20]}_{str(row['created_at'])[:10]}.png",
                        mime="image/png",
                        key=f"dl_map_{row['id']}"
                    )

                # ── Tombol Hapus (selalu muncul, ada map atau tidak) ───────
                st.markdown("---")
                del_confirm = st.checkbox(f"Konfirmasi hapus rekomendasi ini", key=f"del_confirm_{row['id']}")
                if st.button("🗑️ Hapus Rekomendasi", key=f"del_rec_{row['id']}", disabled=not del_confirm):
                    recs_updated = recs[recs['id'] != row['id']]
                    recs_updated.to_excel(RECOMMENDATIONS_FILE, index=False, engine='openpyxl')
                    load_recommendations.clear()
                    st.rerun()

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — DATA MANAGER
# ════════════════════════════════════════════════════════════════════════════
with tab_data:
    st.markdown("## Data Manager")
    po_all = load_po_data()

    if po_all.empty:
        st.markdown('<div class="strip-empty">Belum ada data.</div>', unsafe_allow_html=True)
    else:
        # Summary per PO
        po_summary = (po_all.groupby('po_no')
                      .agg(
                          records=('po_no','count'),
                          area=('area', lambda x: x.mode()[0] if len(x) else ''),
                          recording_date=('recording_date','min'),
                          nnva_count=('time_type', lambda x: (x=='NNVA').sum()),
                          nnva_time=('time_sec', lambda x: x[po_all.loc[x.index,'time_type']=='NNVA'].sum())
                      ).reset_index()
                      .sort_values('recording_date', ascending=False))

        st.markdown(f'<div class="sec-label">{len(po_summary)} PO tersimpan — {len(po_all)} total records</div>',
                    unsafe_allow_html=True)

        # Pilih PO untuk dilihat / dihapus
        selected_po = st.selectbox("Pilih PO untuk dilihat / dihapus",
                                    po_summary['po_no'].tolist(), key="dm_select")

        if selected_po:
            po_detail = po_all[po_all['po_no'] == selected_po]
            d1, d2, d3, d4 = st.columns(4)
            d1.metric("Records", len(po_detail))
            d2.metric("NNVA Records", (po_detail['time_type']=='NNVA').sum())
            d3.metric("NNVA Time", f"{po_detail[po_detail['time_type']=='NNVA']['time_sec'].sum():.0f}s")
            d4.metric("Area", po_detail['area'].mode()[0] if len(po_detail) else '—')

            st.markdown('<div class="sec-label">Detail Data</div>', unsafe_allow_html=True)
            show_cols = [c for c in ['po_no','process_section','time_sec','time_type',
                                      'vz_sec','te_sec','area','item_no','recording_date']
                         if c in po_detail.columns]
            st.dataframe(po_detail[show_cols], use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("### ⚠️ Hapus PO")
            confirm = st.checkbox(f"Saya yakin ingin menghapus semua data PO **{selected_po}**",
                                   key="dm_confirm")
            if st.button("🗑️ Hapus PO Ini", type="primary", disabled=not confirm, key="dm_delete"):
                new_df = po_all[po_all['po_no'] != selected_po]
                if save_po_data(new_df):
                    st.success(f"✓ PO {selected_po} berhasil dihapus ({len(po_detail)} records).")
                    st.rerun()
                else:
                    st.error("Gagal menghapus data.")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("""
<div style='text-align:center;color:#64748b;padding:1.5rem;background:white;
border-radius:0.75rem;font-family:"DM Sans",sans-serif;'>
    <p style='font-weight:600;font-size:1rem;margin-bottom:4px;'>NNVA Analysis System</p>
    <p style='font-size:0.8rem;'>PT Güntner Indonesia · Cycle Time Optimisation</p>
</div>
""", unsafe_allow_html=True)
